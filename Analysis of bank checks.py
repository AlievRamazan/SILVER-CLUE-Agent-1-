import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import re
from datetime import datetime
import os
import sqlite3
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReceiptAnalyzer:
    def __init__(self):
        self.learned_patterns = self.load_patterns()

    def load_patterns(self):
        patterns = {
            'sber': {
                'sender': [
                    r'ФИО отправителя\s*([^\n]+)',
                    r'Отправитель[:\s]*([^\n]+)',
                    r'ФИО[^\n]*отправителя[^\n]*([А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)+)'
                ],
                'receiver': [
                    r'ФИО получателя\s*([^\n]+)',
                    r'Получатель[:\s]*([^\n]+)',
                    r'ФИО[^\n]*получателя[^\n]*([А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)+)'
                ],
                'amount': [
                    r'Сумма перевода\s*([\d\s]+[,\.]\d{2})',
                    r'Сумма[^\d]*([\d\s]+[,\.]\d{2})\s*₽',
                    r'([\d\s]+[,\.]\d{2})\s*₽',
                    r'Перевод[^\d]*([\d\s]+[,\.]\d{2})'
                ],
                'date': [
                    r'(\d{1,2}\s+(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)\s+\d{4})',
                    r'(\d{1,2}\.\d{1,2}\.\d{4})',
                    r'Дата[:\s]*(\d{1,2}\.\d{1,2}\.\d{4})'
                ],
                'phone': [
                    r'Телефон[^\n]*?(\+7[\s\(\-]?\d{3}[\s\)\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2})',
                    r'тел[\.:\s]*([\+7|8][\s\(\-]?\d{3}[\s\)\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2})',
                ],
                'account': [
                    r'Счёт отправителя[^\d]*[\*]*\s*(\d{4})',
                    r'Номер карты получателя[^\d]*[\*]*\s*(\d{4})',
                    r'отправителя[^\d]*[\*]*\s*(\d{4})'
                ]
            }
        }
        return patterns

    def detect_bank(self, text):
        text_lower = text.lower()
        if any(word in text_lower for word in ['сбер', 'sber']):
            return 'sber'
        else:
            return 'sber'

    def extract_entities(self, text):
        bank = self.detect_bank(text)
        patterns = self.learned_patterns.get(bank, {})
        extracted = {'bank': bank}

        for entity_type, pattern_list in patterns.items():
            for pattern in pattern_list:
                matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
                if matches:
                    extracted[entity_type] = matches[0].strip()
                    break

        if 'amount' in extracted:
            amount_str = extracted['amount'].replace(' ', '').replace(',', '.')
            try:
                extracted['amount'] = float(amount_str)
            except ValueError:
                extracted['amount'] = 0.0

        if 'phone' in extracted:
            extracted['phone'] = self.normalize_phone(extracted['phone'])

        if 'date' in extracted:
            extracted['date'] = self.parse_date(extracted['date'])

        if 'sender' in extracted:
            extracted['fio'] = extracted['sender']
        elif 'receiver' in extracted:
            extracted['fio'] = extracted['receiver']

        return extracted

    def normalize_phone(self, phone):
        if not phone:
            return ""
        phone = re.sub(r'\D', '', phone)
        if phone.startswith('+7'):
            phone = '8' + phone[2:]
        elif phone.startswith('7'):
            phone = '8' + phone[1:]
        return phone[:11] if len(phone) >= 11 else phone

    def parse_date(self, date_str):
        try:
            month_map = {
                'января': '01', 'февраля': '02', 'марта': '03',
                'апреля': '04', 'мая': '05', 'июня': '06',
                'июля': '07', 'августа': '08', 'сентября': '09',
                'октября': '10', 'ноября': '11', 'декабря': '12'
            }

            for ru_month, num_month in month_map.items():
                if ru_month in date_str.lower():
                    date_str = date_str.replace(ru_month, num_month)
                    parts = re.findall(r'\d+', date_str)
                    if len(parts) == 3:
                        day, month, year = parts
                        return f"{int(day):02d}.{int(month):02d}.{year}"

            if re.match(r'\d{1,2}\.\d{1,2}\.\d{4}', date_str):
                return date_str

        except Exception as e:
            print(f"Date parsing error: {e}")

        return datetime.now().strftime('%d.%m.%Y')


class AccountingWorkOptimizer:
    def __init__(self):
        self.analyzer = ReceiptAnalyzer()
        self.db_file = "receipts_database.db"
        self.init_database()

    def init_database(self):
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clients (
                client_id INTEGER PRIMARY KEY AUTOINCREMENT,
                fio TEXT NOT NULL,
                phone TEXT,
                account TEXT,
                total_debt REAL DEFAULT 0,
                created_date TEXT
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS payments (
                payment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER,
                amount REAL NOT NULL,
                payment_date TEXT NOT NULL,
                receipt_text TEXT,
                bank_name TEXT,
                created_date TEXT,
                file_hash TEXT,
                is_manual INTEGER DEFAULT 0,
                FOREIGN KEY (client_id) REFERENCES clients (client_id)
            )
        ''')

        try:
            cursor.execute("ALTER TABLE payments ADD COLUMN file_hash TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE payments ADD COLUMN is_manual INTEGER DEFAULT 0")
        except sqlite3.OperationalError:
            pass

        conn.commit()
        conn.close()

    def calculate_file_hash(self, file_path):
        try:
            hasher = hashlib.md5()
            with open(file_path, 'rb') as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hasher.update(chunk)
            return hasher.hexdigest()
        except Exception as e:
            print(f"Hash calculation error: {e}")
            return None

    def is_duplicate_file(self, file_hash):
        if not file_hash:
            return False

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT payment_id FROM payments WHERE file_hash = ?', (file_hash,))
            result = cursor.fetchone()
            conn.close()
            return result is not None
        except Exception as e:
            print(f"Duplicate check error: {e}")
            return False

    def extract_text_from_pdf(self, pdf_path):
        text = ""
        try:
            import PyPDF2
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text
        except ImportError:
            messagebox.showwarning("Warning", "Install PyPDF2: pip install PyPDF2")
            return ""
        except Exception as e:
            print(f"Text extraction error: {e}")
            return ""

    def find_or_create_client(self, fio, phone="", account=""):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            query = "SELECT client_id, total_debt FROM clients WHERE fio = ?"
            params = [fio]

            if phone:
                query += " AND phone = ?"
                params.append(phone)
            if account:
                query += " AND account = ?"
                params.append(account)

            cursor.execute(query, params)
            result = cursor.fetchone()

            if result:
                client_id, total_debt = result
                conn.close()
                return client_id, total_debt
            else:
                total_debt = self.ask_for_debt_info(fio)
                if total_debt is None:
                    conn.close()
                    return None, None

                cursor.execute('''
                    INSERT INTO clients (fio, phone, account, total_debt, created_date)
                    VALUES (?, ?, ?, ?, ?)
                ''', (fio, phone or "", account or "", total_debt, datetime.now().strftime('%d.%m.%Y')))

                client_id = cursor.lastrowid
                conn.commit()
                conn.close()
                return client_id, total_debt

        except Exception as e:
            print(f"Client search/creation error: {e}")
            return None, None

    def add_payment(self, client_id, amount, payment_date, receipt_text, bank_name, file_hash, is_manual=False):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute('''
                INSERT INTO payments (client_id, amount, payment_date, receipt_text, bank_name, created_date, file_hash, is_manual)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (client_id, amount, payment_date, receipt_text, bank_name,
                  datetime.now().strftime('%d.%m.%Y'), file_hash, 1 if is_manual else 0))

            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Payment addition error: {e}")
            return False

    def add_manual_payment(self, client_id, amount, payment_date, description=""):
        return self.add_payment(
            client_id, amount, payment_date,
            f"Manual payment: {description}",
            "Manual entry", "", True
        )

    def delete_payment(self, payment_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM payments WHERE payment_id = ?', (payment_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Payment deletion error: {e}")
            return False

    def delete_client(self, client_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM payments WHERE client_id = ?', (client_id,))
            cursor.execute('DELETE FROM clients WHERE client_id = ?', (client_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Client deletion error: {e}")
            return False

    def update_client(self, client_id, fio=None, phone=None, account=None, total_debt=None):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            updates = []
            params = []

            if fio is not None:
                updates.append("fio = ?")
                params.append(fio)
            if phone is not None:
                updates.append("phone = ?")
                params.append(phone)
            if account is not None:
                updates.append("account = ?")
                params.append(account)
            if total_debt is not None:
                updates.append("total_debt = ?")
                params.append(total_debt)

            if updates:
                params.append(client_id)
                cursor.execute(f'UPDATE clients SET {", ".join(updates)} WHERE client_id = ?', params)

            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Client update error: {e}")
            return False

    def apply_discount(self, client_id, discount_amount):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute('SELECT total_debt FROM clients WHERE client_id = ?', (client_id,))
            result = cursor.fetchone()
            if not result:
                return None

            current_debt = result[0]
            new_debt = max(0, current_debt - discount_amount)
            cursor.execute('UPDATE clients SET total_debt = ? WHERE client_id = ?', (new_debt, client_id))

            conn.commit()
            conn.close()
            return new_debt
        except Exception as e:
            print(f"Discount application error: {e}")
            return None

    def get_client_info(self, client_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM clients WHERE client_id = ?', (client_id,))
            result = cursor.fetchone()
            conn.close()
            return result
        except Exception as e:
            print(f"Client info retrieval error: {e}")
            return None

    def get_all_clients(self):
        try:
            conn = sqlite3.connect(self.db_file)
            clients_df = pd.read_sql('SELECT * FROM clients ORDER BY fio', conn)
            conn.close()
            return clients_df
        except Exception as e:
            print(f"Clients retrieval error: {e}")
            return pd.DataFrame()

    def get_all_payments(self):
        try:
            conn = sqlite3.connect(self.db_file)
            payments_df = pd.read_sql('''
                SELECT p.*, c.fio 
                FROM payments p 
                LEFT JOIN clients c ON p.client_id = c.client_id 
                ORDER BY p.payment_date DESC
            ''', conn)
            conn.close()
            return payments_df
        except Exception as e:
            print(f"Payments retrieval error: {e}")
            return pd.DataFrame()

    def calculate_remaining_debt(self, client_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute('SELECT total_debt FROM clients WHERE client_id = ?', (client_id,))
            result = cursor.fetchone()
            if not result:
                return 0

            total_debt = result[0]

            cursor.execute('SELECT SUM(amount) FROM payments WHERE client_id = ?', (client_id,))
            total_payments_result = cursor.fetchone()
            total_payments = total_payments_result[0] if total_payments_result[0] is not None else 0

            conn.close()
            return total_debt - total_payments
        except Exception as e:
            print(f"Debt calculation error: {e}")
            return 0

    def ask_for_debt_info(self, fio):
        try:
            response = messagebox.askyesno(
                "New Client",
                f"New client detected: {fio}\n\nAdd to database?"
            )

            if response:
                total_debt = simpledialog.askfloat(
                    "Total Debt",
                    f"Enter total debt for {fio}:",
                    initialvalue=1000.0,
                    minvalue=0.0
                )
                return total_debt
            return None
        except Exception as e:
            print(f"Debt info request error: {e}")
            return None

    def process_receipt(self, text, filename, file_hash):
        try:
            if self.is_duplicate_file(file_hash):
                return f"⏭️ {filename}: Skipped (already processed)"

            extracted_data = self.analyzer.extract_entities(text)
            if not extracted_data:
                return f"❌ {filename}: Failed to recognize receipt data"

            if 'fio' not in extracted_data:
                return f"❌ {filename}: Failed to determine name"

            if 'amount' not in extracted_data or extracted_data['amount'] <= 0:
                return f"❌ {filename}: Failed to determine amount"

            client_id, total_debt = self.find_or_create_client(
                extracted_data['fio'],
                extracted_data.get('phone', ''),
                extracted_data.get('account', '')
            )

            if client_id is None:
                return f"⏸️ {filename}: Skipped - {extracted_data['fio']}"

            success = self.add_payment(
                client_id,
                extracted_data['amount'],
                extracted_data.get('date', datetime.now().strftime('%d.%m.%Y')),
                text[:500],
                extracted_data['bank'],
                file_hash
            )

            if not success:
                return f"❌ {filename}: Payment save error"

            remaining_debt = self.calculate_remaining_debt(client_id)

            return f"✅ {extracted_data['fio']}: payment {extracted_data['amount']} rub. (remaining: {remaining_debt:.2f} rub.)"

        except Exception as e:
            print(f"Receipt processing critical error: {e}")
            return f"❌ {filename}: Processing error - {str(e)}"

    def process_pdf_files(self, pdf_files):
        results = []

        for pdf_file in pdf_files:
            file_hash = self.calculate_file_hash(pdf_file)
            if not file_hash:
                results.append(f"❌ {os.path.basename(pdf_file)}: file read error")
                continue

            text = self.extract_text_from_pdf(pdf_file)
            if not text.strip():
                results.append(f"❌ {os.path.basename(pdf_file)}: failed to extract text")
                continue

            result = self.process_receipt(text, os.path.basename(pdf_file), file_hash)
            results.append(result)

        return results

    def get_database_stats(self):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute('SELECT COUNT(*) FROM clients')
            total_clients = cursor.fetchone()[0]

            cursor.execute('SELECT COUNT(*) FROM payments')
            total_payments = cursor.fetchone()[0]

            cursor.execute('SELECT SUM(amount) FROM payments')
            total_amount_result = cursor.fetchone()
            total_amount = total_amount_result[0] if total_amount_result[0] is not None else 0

            conn.close()
            return total_clients, total_payments, total_amount
        except Exception as e:
            print(f"Statistics retrieval error: {e}")
            return 0, 0, 0

    def export_to_excel(self):
        try:
            clients_df = self.get_all_clients()
            payments_df = self.get_all_payments()

            if clients_df.empty and payments_df.empty:
                messagebox.showinfo("Information", "No data to export")
                return False

            if not clients_df.empty:
                clients_export = clients_df.copy()
                clients_export['Paid'] = clients_export['client_id'].apply(
                    lambda x: self.get_total_payments(x)
                )
                clients_export['Remaining_Debt'] = clients_export['client_id'].apply(
                    lambda x: self.calculate_remaining_debt(x)
                )
                clients_export['Payment_Count'] = clients_export['client_id'].apply(
                    lambda x: self.get_payment_count(x)
                )
            else:
                clients_export = pd.DataFrame()

            file_path = filedialog.asksaveasfilename(
                title="Save Excel Report",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )

            if file_path:
                self.create_beautiful_excel(file_path, clients_export, payments_df)
                return True
            return False

        except Exception as e:
            print(f"Excel export error: {e}")
            messagebox.showerror("Error", f"Export error: {str(e)}")
            return False

    def get_total_payments(self, client_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(amount) FROM payments WHERE client_id = ?', (client_id,))
            result = cursor.fetchone()
            conn.close()
            return result[0] if result[0] is not None else 0
        except:
            return 0

    def get_payment_count(self, client_id):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM payments WHERE client_id = ?', (client_id,))
            result = cursor.fetchone()
            conn.close()
            return result[0] if result[0] is not None else 0
        except:
            return 0

    def create_beautiful_excel(self, file_path, clients_df, payments_df):
        try:
            wb = Workbook()

            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            money_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            money_format = '#,##0.00" rub."'
            date_format = 'DD.MM.YYYY'

            if not clients_df.empty:
                ws_clients = wb.active
                ws_clients.title = "Clients"

                headers = ["ID", "Name", "Phone", "Account", "Total Debt", "Paid", "Remaining Debt",
                           "Payment Count", "Date Added"]
                for col, header in enumerate(headers, 1):
                    cell = ws_clients.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align

                for row, (_, client_row) in enumerate(clients_df.iterrows(), 2):
                    for col, value in enumerate([client_row['client_id'], client_row['fio'],
                                                 client_row['phone'], client_row['account'],
                                                 client_row['total_debt'], client_row['Paid'],
                                                 client_row['Remaining_Debt'], client_row['Payment_Count'],
                                                 client_row['created_date']], 1):
                        cell = ws_clients.cell(row=row, column=col, value=value)
                        cell.border = border

                        if col in [5, 6, 7]:
                            if pd.notna(value):
                                cell.number_format = money_format
                                cell.fill = money_fill
                            cell.alignment = center_align
                        elif col == 9:
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align

                for column in ws_clients.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws_clients.column_dimensions[column_letter].width = adjusted_width

                ws_clients.freeze_panes = 'A2'

            if not payments_df.empty:
                ws_payments = wb.create_sheet("Payment History")

                headers = ["ID", "Name", "Amount", "Payment Date", "Bank", "Type", "Date Added"]
                for col, header in enumerate(headers, 1):
                    cell = ws_payments.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align

                for row, (_, payment_row) in enumerate(payments_df.iterrows(), 2):
                    payment_type = "Manual" if payment_row['is_manual'] == 1 else "Auto"

                    for col, value in enumerate([payment_row['payment_id'], payment_row['fio'],
                                                 payment_row['amount'], payment_row['payment_date'],
                                                 payment_row['bank_name'], payment_type,
                                                 payment_row['created_date']], 1):
                        cell = ws_payments.cell(row=row, column=col, value=value)
                        cell.border = border

                        if col == 3:
                            if pd.notna(value):
                                cell.number_format = money_format
                                cell.fill = money_fill
                            cell.alignment = center_align
                        elif col in [4, 7]:
                            cell.alignment = center_align
                        elif col == 6:
                            if value == 'Manual':
                                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align

                for column in ws_payments.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    ws_payments.column_dimensions[column_letter].width = adjusted_width

                ws_payments.freeze_panes = 'A2'

            wb.save(file_path)
            print(f"Excel file saved: {file_path}")

        except Exception as e:
            print(f"Excel creation error: {e}")
            raise


class AccountingOptimizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Accounting Work Optimizer")
        self.root.geometry("1000x700")

        self.optimizer = AccountingWorkOptimizer()
        self.setup_ui()
        self.update_stats()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(main_frame,
                                text="Accounting Work Optimizer",
                                font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))

        self.stats_label = ttk.Label(main_frame,
                                     text="Loading statistics...",
                                     font=('Arial', 11),
                                     relief='solid',
                                     padding=10)
        self.stats_label.pack(fill=tk.X, pady=(0, 20))

        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=(0, 20))

        row1 = ttk.Frame(buttons_frame)
        row1.pack(fill=tk.X, pady=5)

        ttk.Button(row1, text="📎 Analyze Receipts",
                   command=self.process_files).pack(side=tk.LEFT, padx=5)

        ttk.Button(row1, text="📊 Export to Excel",
                   command=self.export_excel).pack(side=tk.LEFT, padx=5)

        ttk.Button(row1, text="👥 Manage Clients",
                   command=self.manage_clients).pack(side=tk.LEFT, padx=5)

        row2 = ttk.Frame(buttons_frame)
        row2.pack(fill=tk.X, pady=5)

        ttk.Button(row2, text="➕ Manual Payment",
                   command=self.add_manual_payment).pack(side=tk.LEFT, padx=5)

        ttk.Button(row2, text="🎁 Apply Discount",
                   command=self.apply_discount).pack(side=tk.LEFT, padx=5)

        ttk.Button(row2, text="🗑️ Manage Payments",
                   command=self.manage_payments).pack(side=tk.LEFT, padx=5)

        ttk.Button(row2, text="🔄 Update Statistics",
                   command=self.update_stats).pack(side=tk.LEFT, padx=5)

        info_text = """
🎯 Bank Receipt Analysis System

• 📎 Analyze Receipts - process PDF receipt files
• 📊 Export to Excel - save data in formatted Excel
• 👥 Manage Clients - add, edit, delete clients
• ➕ Manual Payment - add payment without receipt
• 🎁 Apply Discount - reduce debt amount
• 🗑️ Manage Payments - view and delete payments

💡 Required for PDF processing:
   pip install PyPDF2
        """

        info_label = ttk.Label(main_frame, text=info_text,
                               justify=tk.LEFT,
                               font=('Arial', 9))
        info_label.pack(fill=tk.BOTH, expand=True)

    def update_stats(self):
        try:
            total_clients, total_payments, total_amount = self.optimizer.get_database_stats()
            stats_text = f"👥 Clients: {total_clients} | 💰 Payments: {total_payments} | 💵 Amount: {total_amount:,.2f} rub."
            self.stats_label.config(text=stats_text)
        except Exception as e:
            print(f"Statistics update error: {e}")

    def process_files(self):
        try:
            pdf_files = filedialog.askopenfilenames(
                title="Select PDF Receipt Files",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )

            if pdf_files:
                results = self.optimizer.process_pdf_files(pdf_files)
                result_text = "\n".join(results)
                messagebox.showinfo("Processing Results", result_text)
                self.update_stats()
        except Exception as e:
            messagebox.showerror("Error", f"File processing error: {str(e)}")

    def export_excel(self):
        try:
            if self.optimizer.export_to_excel():
                messagebox.showinfo("Success", "Data exported to Excel")
        except Exception as e:
            messagebox.showerror("Error", f"Export error: {str(e)}")

    def manage_clients(self):
        try:
            clients_df = self.optimizer.get_all_clients()

            if clients_df.empty:
                messagebox.showinfo("Clients", "No clients in database")
                return

            window = tk.Toplevel(self.root)
            window.title("Manage Clients")
            window.geometry("900x600")

            tree = ttk.Treeview(window, columns=("ID", "Name", "Phone", "Account", "Debt"), show="headings")
            tree.heading("ID", text="ID")
            tree.heading("Name", text="Name")
            tree.heading("Phone", text="Phone")
            tree.heading("Account", text="Account")
            tree.heading("Debt", text="Debt")

            tree.column("ID", width=50)
            tree.column("Name", width=200)
            tree.column("Phone", width=150)
            tree.column("Account", width=100)
            tree.column("Debt", width=100)

            for _, client in clients_df.iterrows():
                tree.insert("", "end", values=(
                    client['client_id'], client['fio'], client['phone'] or "",
                    client['account'] or "", f"{client['total_debt']:.2f} rub."
                ))

            tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            button_frame = ttk.Frame(window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)

            ttk.Button(button_frame, text="✏️ Edit",
                       command=lambda: self.edit_client(tree, window)).pack(side=tk.LEFT, padx=5)

            ttk.Button(button_frame, text="🗑️ Delete",
                       command=lambda: self.delete_client(tree, window)).pack(side=tk.LEFT, padx=5)

            ttk.Button(button_frame, text="➕ Add Payment",
                       command=lambda: self.add_payment_to_client(tree, window)).pack(side=tk.LEFT, padx=5)

            ttk.Button(button_frame, text="🎁 Discount",
                       command=lambda: self.apply_discount_to_client(tree, window)).pack(side=tk.LEFT, padx=5)

            ttk.Button(button_frame, text="Close",
                       command=window.destroy).pack(side=tk.RIGHT, padx=5)

        except Exception as e:
            messagebox.showerror("Error", f"Client loading error: {str(e)}")

    def edit_client(self, tree, window):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Select client to edit")
            return

        item = selected[0]
        values = tree.item(item, 'values')
        client_id = values[0]

        client_info = self.optimizer.get_client_info(client_id)
        if not client_info:
            messagebox.showerror("Error", "Failed to load client data")
            return

        edit_window = tk.Toplevel(window)
        edit_window.title("Edit Client")
        edit_window.geometry("400x300")

        ttk.Label(edit_window, text="Name:").pack(pady=5)
        fio_entry = ttk.Entry(edit_window, width=50)
        fio_entry.insert(0, client_info[1])
        fio_entry.pack(pady=5)

        ttk.Label(edit_window, text="Phone:").pack(pady=5)
        phone_entry = ttk.Entry(edit_window, width=50)
        phone_entry.insert(0, client_info[2] or "")
        phone_entry.pack(pady=5)

        ttk.Label(edit_window, text="Account:").pack(pady=5)
        account_entry = ttk.Entry(edit_window, width=50)
        account_entry.insert(0, client_info[3] or "")
        account_entry.pack(pady=5)

        ttk.Label(edit_window, text="Total Debt:").pack(pady=5)
        debt_entry = ttk.Entry(edit_window, width=50)
        debt_entry.insert(0, str(client_info[4]))
        debt_entry.pack(pady=5)

        def save_changes():
            new_fio = fio_entry.get()
            new_phone = phone_entry.get()
            new_account = account_entry.get()
            try:
                new_debt = float(debt_entry.get())
            except ValueError:
                messagebox.showerror("Error", "Enter valid debt amount")
                return

            success = self.optimizer.update_client(client_id, new_fio, new_phone, new_account, new_debt)
            if success:
                messagebox.showinfo("Success", "Client data updated")
                edit_window.destroy()
                window.destroy()
                self.manage_clients()
                self.update_stats()
            else:
                messagebox.showerror("Error", "Failed to update client data")

        ttk.Button(edit_window, text="💾 Save",
                   command=save_changes).pack(pady=10)

    def delete_client(self, tree, window):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Select client to delete")
            return

        item = selected[0]
        values = tree.item(item, 'values')
        client_id = values[0]
        client_name = values[1]

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Delete client {client_name}?\n\nThis will delete all related payments!"
        )

        if confirm:
            success = self.optimizer.delete_client(client_id)
            if success:
                messagebox.showinfo("Success", "Client deleted")
                window.destroy()
                self.manage_clients()
                self.update_stats()
            else:
                messagebox.showerror("Error", "Failed to delete client")

    def add_payment_to_client(self, tree, window):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Select client for payment")
            return

        item = selected[0]
        values = tree.item(item, 'values')
        client_id = values[0]
        client_name = values[1]

        amount = simpledialog.askfloat("Payment Amount",
                                       f"Enter payment amount for {client_name}:",
                                       minvalue=0.01)
        if amount is None:
            return

        description = simpledialog.askstring("Description", "Enter payment description (optional):")

        success = self.optimizer.add_manual_payment(
            client_id, amount,
            datetime.now().strftime('%d.%m.%Y'),
            description or ""
        )

        if success:
            messagebox.showinfo("Success", f"Payment {amount} rub. added for {client_name}")
            window.destroy()
            self.manage_clients()
            self.update_stats()
        else:
            messagebox.showerror("Error", "Failed to add payment")

    def apply_discount_to_client(self, tree, window):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Select client for discount")
            return

        item = selected[0]
        values = tree.item(item, 'values')
        client_id = values[0]
        client_name = values[1]

        current_debt = self.optimizer.calculate_remaining_debt(client_id)

        discount = simpledialog.askfloat("Discount Amount",
                                         f"Current debt {client_name}: {current_debt:.2f} rub.\n\nEnter discount amount:",
                                         minvalue=0.01, maxvalue=current_debt)
        if discount is None:
            return

        new_debt = self.optimizer.apply_discount(client_id, discount)
        if new_debt is not None:
            messagebox.showinfo("Success",
                                f"Discount applied!\n\nClient: {client_name}\nDiscount: {discount:.2f} rub.\nNew debt: {new_debt:.2f} rub.")
            window.destroy()
            self.manage_clients()
            self.update_stats()
        else:
            messagebox.showerror("Error", "Failed to apply discount")

    def add_manual_payment(self):
        try:
            clients_df = self.optimizer.get_all_clients()

            if clients_df.empty:
                messagebox.showinfo("Clients", "No clients in database")
                return

            payment_window = tk.Toplevel(self.root)
            payment_window.title("Manual Payment")
            payment_window.geometry("400x200")

            ttk.Label(payment_window, text="Select client:").pack(pady=10)

            client_var = tk.StringVar()
            client_combo = ttk.Combobox(payment_window, textvariable=client_var, width=50)
            client_combo['values'] = [f"{row['fio']} (ID: {row['client_id']})" for _, row in clients_df.iterrows()]
            client_combo.pack(pady=10)

            ttk.Label(payment_window, text="Payment amount:").pack(pady=10)
            amount_entry = ttk.Entry(payment_window, width=50)
            amount_entry.pack(pady=10)

            def process_payment():
                client_str = client_var.get()
                if not client_str:
                    messagebox.showwarning("Error", "Select client")
                    return

                try:
                    amount = float(amount_entry.get())
                except ValueError:
                    messagebox.showwarning("Error", "Enter valid amount")
                    return

                client_id = int(client_str.split("(ID: ")[1].replace(")", ""))

                success = self.optimizer.add_manual_payment(
                    client_id, amount,
                    datetime.now().strftime('%d.%m.%Y'),
                    "Manual payment"
                )

                if success:
                    messagebox.showinfo("Success", f"Manual payment {amount} rub. added")
                    payment_window.destroy()
                    self.update_stats()
                else:
                    messagebox.showerror("Error", "Failed to add payment")

            ttk.Button(payment_window, text="💾 Add Payment",
                       command=process_payment).pack(pady=20)

        except Exception as e:
            messagebox.showerror("Error", f"Payment addition error: {str(e)}")

    def apply_discount(self):
        try:
            clients_df = self.optimizer.get_all_clients()

            if clients_df.empty:
                messagebox.showinfo("Clients", "No clients in database")
                return

            discount_window = tk.Toplevel(self.root)
            discount_window.title("Apply Discount")
            discount_window.geometry("500x300")

            ttk.Label(discount_window, text="Select client:").pack(pady=10)

            client_var = tk.StringVar()
            client_combo = ttk.Combobox(discount_window, textvariable=client_var, width=50)

            client_list = []
            self.client_debts = {}
            for _, row in clients_df.iterrows():
                current_debt = self.optimizer.calculate_remaining_debt(row['client_id'])
                client_str = f"{row['fio']} (Debt: {current_debt:.2f} rub.)"
                client_list.append(client_str)
                self.client_debts[client_str] = (row['client_id'], current_debt)

            client_combo['values'] = client_list
            client_combo.pack(pady=10)

            debt_label = ttk.Label(discount_window, text="Current debt: -", font=('Arial', 10, 'bold'))
            debt_label.pack(pady=10)

            def update_debt_label(event):
                client_str = client_var.get()
                if client_str in self.client_debts:
                    _, current_debt = self.client_debts[client_str]
                    debt_label.config(text=f"Current debt: {current_debt:.2f} rub.")

            client_combo.bind('<<ComboboxSelected>>', update_debt_label)

            ttk.Label(discount_window, text="Discount amount:").pack(pady=10)
            discount_entry = ttk.Entry(discount_window, width=50)
            discount_entry.pack(pady=10)

            def process_discount():
                client_str = client_var.get()
                if not client_str:
                    messagebox.showwarning("Error", "Select client")
                    return

                try:
                    discount = float(discount_entry.get())
                except ValueError:
                    messagebox.showwarning("Error", "Enter valid discount amount")
                    return

                client_id, current_debt = self.client_debts[client_str]

                if discount > current_debt:
                    messagebox.showwarning("Error", "Discount cannot exceed current debt")
                    return

                new_debt = self.optimizer.apply_discount(client_id, discount)
                if new_debt is not None:
                    messagebox.showinfo("Success",
                                        f"Discount applied!\n\nClient: {client_str.split(' (')[0]}\nDiscount: {discount:.2f} rub.\nNew debt: {new_debt:.2f} rub.")
                    discount_window.destroy()
                    self.update_stats()
                else:
                    messagebox.showerror("Error", "Failed to apply discount")

            ttk.Button(discount_window, text="🎁 Apply Discount",
                       command=process_discount).pack(pady=20)

        except Exception as e:
            messagebox.showerror("Error", f"Discount application error: {str(e)}")

    def manage_payments(self):
        try:
            payments_df = self.optimizer.get_all_payments()

            if payments_df.empty:
                messagebox.showinfo("Payments", "No payments in database")
                return

            window = tk.Toplevel(self.root)
            window.title("Manage Payments")
            window.geometry("1000x600")

            tree = ttk.Treeview(window, columns=("ID", "Name", "Amount", "Date", "Bank", "Type"), show="headings")
            tree.heading("ID", text="ID")
            tree.heading("Name", text="Name")
            tree.heading("Amount", text="Amount")
            tree.heading("Date", text="Date")
            tree.heading("Bank", text="Bank")
            tree.heading("Type", text="Type")

            tree.column("ID", width=50)
            tree.column("Name", width=200)
            tree.column("Amount", width=100)
            tree.column("Date", width=100)
            tree.column("Bank", width=150)
            tree.column("Type", width=100)

            for _, payment in payments_df.iterrows():
                payment_type = "Manual" if payment['is_manual'] == 1 else "Auto"
                tree.insert("", "end", values=(
                    payment['payment_id'], payment['fio'] or "Unknown",
                    f"{payment['amount']:.2f} rub.", payment['payment_date'],
                    payment['bank_name'] or "", payment_type
                ))

            tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            button_frame = ttk.Frame(window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)

            ttk.Button(button_frame, text="🗑️ Delete",
                       command=lambda: self.delete_payment(tree, window)).pack(side=tk.LEFT, padx=5)

            ttk.Button(button_frame, text="Close",
                       command=window.destroy).pack(side=tk.RIGHT, padx=5)

        except Exception as e:
            messagebox.showerror("Error", f"Payments loading error: {str(e)}")

    def delete_payment(self, tree, window):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Select payment to delete")
            return

        item = selected[0]
        values = tree.item(item, 'values')
        payment_id = values[0]

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Delete this payment?"
        )

        if confirm:
            success = self.optimizer.delete_payment(payment_id)
            if success:
                messagebox.showinfo("Success", "Payment deleted")
                window.destroy()
                self.manage_payments()
                self.update_stats()
            else:
                messagebox.showerror("Error", "Failed to delete payment")


if __name__ == "__main__":
    try:
        import pandas as pd
        import sqlite3

        print("✅ All required libraries installed")
    except ImportError as e:
        print(f"❌ Missing library: {e}")
        print("Install: pip install pandas")

    root = tk.Tk()
    app = AccountingOptimizerApp(root)
    root.mainloop()